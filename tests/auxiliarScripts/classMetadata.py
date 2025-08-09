import openpyxl

# Carrega a planilha e a aba especificada
workbook = openpyxl.load_workbook('dataImport.xlsx')
# sheetName = 'classMetadata'
sheetName = 'propsMetadata'
sheet = workbook[sheetName]

outputFilename = sheetName + '.ttl'

# Abre o arquivo de saída
with open(outputFilename, 'w', encoding='utf-8') as file:
    # Pula a primeira linha (header) e começa da segunda
    for n, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # Verifica se a célula A está vazia (condição de parada)
        if row[0].value is None:
            break
        
        objectName = ':' + str(row[0].value)  # : + célula An
        
        # Loop das colunas A até H (índices 0 a 7)
        for col_idx in range(8):  # A=0, B=1, ..., H=7
            header = sheet.cell(row=1, column=col_idx + 1).value  # Pega o header da coluna (linha 1)
            cell_input = row[col_idx].value

            # Se a célula estiver vazia, pula para a próxima coluna
            if cell_input is None or str(cell_input).strip() == "":
                continue
            
            # Formata a string conforme o idioma do header
            if '"' in cell_input: cell_input = cell_input.replace('"','\"')
            if header.endswith("@en"):
                h = header.replace(" @en", "")
                line = f"{objectName} {h} \"{cell_input}\"@en .\n"
            elif header.endswith("@pt-BR"):
                h = header.replace(" @pt-BR", "")
                line = f"{objectName} {h} \"{cell_input}\"@pt-BR .\n"
            else:
                if cell_input == 'IfcPropertySingleValue':
                    cell_value = ':singleValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcTable':
                    continue
                elif cell_input == 'IfcPropertyListValue':
                    cell_value = ':multiValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcInteger':
                    cell_value = ':intValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcReal':
                    cell_value = ':floatValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcText':
                    cell_value = ':stringValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcBoolean':
                    cell_value = ':booleanValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                elif cell_input == 'IfcTimeStamp':
                    cell_value = ':dateTimeValue'
                    line = f"{objectName} {header} {cell_value} .\n"
                else:
                    cell_value = cell_input
                    line = f"{objectName} {header} \"{cell_value}\" .\n"
            
            file.write(line)
        
        # Verifica a célula I (índice 8) para skos:altLabel
        if row[8].value is not None:  # Se a célula In não estiver vazia
            header = sheet.cell(row=1, column=9).value  # Pega o header da coluna (linha 1)
            parts = str(row[8].value).split('; ')  # Split pelo separador
            if header == ':validValues':
                continue
            else:
                for part in parts:
                    if part.strip():  # Só escreve se a parte não for vazia
                        alt_label_line = f"{objectName} skos:altLabel \"{part}\"@en .\n"
                        file.write(alt_label_line)

print("Processamento concluído. Resultados salvos em " + outputFilename + ".")